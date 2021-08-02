import os
import re
import sys
from urllib.request import urlopen, Request
from datetime import datetime
import requests

from openpyxl import Workbook
from bs4 import BeautifulSoup

print("Press 1, for web scraping using URL")
print("Press 2, for web scraping using Keywords")
number = int(input("Enter 1 for URL and 2 for Keyword : "))

save_excel = True 

book = Workbook()
sheet1 = book.create_sheet("E-mails",0)
sheet2 = book.create_sheet("Phone_numbers",1)

if number==1:
    def start_scrape(page, name_the_file):
        print("\n\nWebpage is currently being scrapped... please wait...")
       
        scrape = BeautifulSoup(page, 'html.parser')
        scrape = scrape.get_text()
    
        phone_numbers = set(re.findall(r"((?:\d{3}|\(\d{3}\))?(?:\s|-|\.)?\d{3}(?:\s|-|\.)\d{4})", scrape)) 
        emails = set(re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,3}", scrape))

        nodupnumber = len(list(phone_numbers))
        nodupemail = len(list(emails))

        dupnumber = len(list(re.findall(r"((?:\d{3}|\(\d{3}\))?(?:\s|-|\.)?\d{3}(?:\s|-|\.)\d{4})", scrape))) 
        dupemail = len(list(re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,3}", scrape)))

        number_of_dup_number = int(dupnumber) - int(nodupnumber) 
        number_of_dup_email = int(dupemail) - int(nodupemail)

        email_list = list(emails)
        phone_numbers_list = list(phone_numbers)

        if len(phone_numbers) == 0:
            print("No phone number(s) found.")
            print("-----------------------------\n")
        else:
            count = 1
            for item in phone_numbers:
                print("Phone number #" + str(count) + ': ' + item)
                count += 1
                print("-----------------------------\n")

        if len(emails) == 0:
            print("No email address(es) found.")
            print("-----------------------------\n")
        else:
            count = 1
            for item in emails:
                print('Email address #' + str(count) + ': ' + item)
                count += 1

        if save_excel:
            for row in zip(email_list):
                sheet1.append(row)
            for num in zip(phone_numbers_list):
                sheet2.append(num)
            excel_file = (name_the_file + ".xlsx")
            book.save(excel_file) 
       
        print("\nDuplicates have been removed from list.")
        print("Total phone numbers: ", nodupnumber)
        print("Total email addresses: ", nodupemail)
        print("There were " + str(number_of_dup_number) + " duplicate phone numbers.")
        print("There were " + str(number_of_dup_email) + " duplicate email addresses.")

        if save_excel:
            print("\n\nData has been stored inside of an Excel spreadsheet named: "
                  + excel_file + " in this directory: " + os.getcwd())
            mod_time = os.stat(excel_file).st_mtime
            print("\nCompleted at: " + str(datetime.fromtimestamp(mod_time)))
            print("\nSize of file: " + str(os.stat(excel_file).st_size) + " KB")
        
    webpage = input("Paste the webpage you would like to scrape (include http/https): ")
    
    name_the_file = input("Name the file you would like to save the data in (don't include .xlsx): ")
    try:
        page = urlopen(webpage) 
        start_scrape(page)
    except:
        hdr = {'User-Agent': 'Mozilla/5.0'}
        req = Request(webpage, headers=hdr)
        page = urlopen(req)
        start_scrape(page, name_the_file)
        
if number==2:
    def start_scrape(page, name_the_file):
        print("\n\nWebpage is currently being scrapped... please wait...")
       
        scrape = BeautifulSoup(page, 'html.parser')
        scrape = scrape.get_text()
    
        phone_numbers = set(re.findall(r"((?:\d{3}|\(\d{3}\))?(?:\s|-|\.)?\d{3}(?:\s|-|\.)\d{4})", scrape)) 
        emails = set(re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,3}", scrape))

        nodupnumber = len(list(phone_numbers))
        nodupemail = len(list(emails))

        dupnumber = len(list(re.findall(r"((?:\d{3}|\(\d{3}\))?(?:\s|-|\.)?\d{3}(?:\s|-|\.)\d{4})", scrape))) 
        dupemail = len(list(re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,3}", scrape)))

        number_of_dup_number = int(dupnumber) - int(nodupnumber) 
        number_of_dup_email = int(dupemail) - int(nodupemail)

        email_list = list(emails)
        phone_numbers_list = list(phone_numbers)

        if len(phone_numbers) == 0:
            print("No phone number(s) found.")
            print("-----------------------------\n")
        else:
            count = 1
            for item in phone_numbers:
                print("Phone number #" + str(count) + ': ' + item)
                count += 1
                print("-----------------------------\n")

        if len(emails) == 0:
            print("No email address(es) found.")
            print("-----------------------------\n")
        else:
            count = 1
            for item in emails:
                print('Email address #' + str(count) + ': ' + item)
                count += 1

        if save_excel:
            for row in zip(email_list):
                sheet1.append(row)
            for num in zip(phone_numbers_list):
                sheet2.append(num)
            excel_file = (name_the_file + ".xlsx")
            book.save(excel_file) 
       
        print("\nDuplicates have been removed from list.")
        print("Total phone numbers: ", nodupnumber)
        print("Total email addresses: ", nodupemail)
        print("There were " + str(number_of_dup_number) + " duplicate phone numbers.")
        print("There were " + str(number_of_dup_email) + " duplicate email addresses.")

        if save_excel:
            print("\n\nData has been stored inside of an Excel spreadsheet named: "
                  + excel_file + " in this directory: " + os.getcwd())
            mod_time = os.stat(excel_file).st_mtime
            print("\nCompleted at: " + str(datetime.fromtimestamp(mod_time)))
            print("\nSize of file: " + str(os.stat(excel_file).st_size) + " KB")

    keyword = str(input("Enter a keyword you wanna search(without any space: )"))        
    res ='https://www.google.com/search?as_q='+''.join(keyword)
    name_the_file = input("Name the file you would like to save the data in (don't include .xlsx): ")
    try:
        page = urlopen(webpage) 
        start_scrape(page)
    except:
        hdr = {'User-Agent': 'Mozilla/5.0'}
        req = Request(res, headers=hdr)
        page = urlopen(req)
        start_scrape(page, name_the_file)
    



    


    
    
